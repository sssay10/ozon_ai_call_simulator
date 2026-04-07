"""Parse «РсВ ФАК» sheet from corp_data.xlsx into Q&A documents."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass(frozen=True)
class FAQDocument:
    """One FAQ block for embedding."""

    question: str
    answer: str
    sheet: str
    row_start: int

    def as_embedding_text(self) -> str:
        return f"Вопрос: {self.question}\n\nОтвет: {self.answer}"


def _cell(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).replace("\xa0", " ").strip()
    return " ".join(s.split())


def parse_rsv_faq_sheet(xlsx_path: Path, sheet_name: str = "РсВ ФАК") -> list[FAQDocument]:
    """Fold multi-row Q/A blocks into FAQDocument rows."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet {sheet_name!r} not in {xlsx_path}, have {wb.sheetnames}")
    ws = wb[sheet_name]
    rows: list[tuple[int, str, str]] = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx == 1:
            continue  # header
        qc = row[0] if len(row) > 0 else None
        ac = row[1] if len(row) > 1 else None
        rows.append((idx, _cell(qc), _cell(ac)))
    wb.close()

    docs: list[FAQDocument] = []
    pending_q: str | None = None
    pending_a: list[str] = []
    row_start = 2

    def flush() -> None:
        nonlocal pending_q, pending_a, row_start
        if not pending_q:
            pending_a.clear()
            return
        ans = "\n".join(x for x in pending_a if x).strip()
        if not ans:
            pending_q = None
            pending_a.clear()
            return
        docs.append(
            FAQDocument(
                question=pending_q.strip(),
                answer=ans,
                sheet=sheet_name,
                row_start=row_start,
            )
        )
        pending_q = None
        pending_a.clear()

    for idx, q, a in rows:
        if not q and not a:
            continue
        if q and a:
            if pending_q and not pending_a:
                # orphan question without answer in sheet — drop when a real block starts
                pending_q = None
            else:
                flush()
            pending_q = q
            pending_a = [a]
            row_start = idx
        elif q and not a:
            if pending_q and pending_a:
                flush()
            pending_q = (pending_q + "\n" + q).strip() if pending_q else q
            if not pending_a:
                row_start = idx
        else:
            if pending_q is None:
                continue
            pending_a.append(a)

    flush()
    return docs
