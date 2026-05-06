"""Ingest bank_docs.xlsx into ChromaDB for KNOWLEDGE_RAG=1 eval mode.

Reads two sheets:
  - РсВ ФАК      : FAQ Q&A pairs, groups multi-row answers into one chunk.
  - Параметры РКО: tariff parameters, one chunk per tariff column.

Recreates the 'rko_faq' collection on each run (idempotent).

Usage:
    python scripts/ingest_faq.py

Environment variables:
    CHROMA_HTTP_HOST  default: localhost
    CHROMA_HTTP_PORT  default: 8000
"""

from __future__ import annotations

import os
import sys
from pathlib import Path

import chromadb
import openpyxl

XLSX_PATH = Path(__file__).resolve().parents[1] / "knowledge_base" / "raw" / "bank_docs.xlsx"
COLLECTION_NAME = "rko_faq"
FAQ_SHEET = "РсВ ФАК"
TARIFF_SHEET = "Параметры РКО"


def _clean(value: object) -> str:
    return str(value).replace("\xa0", " ").strip() if value else ""


def parse_faq(ws) -> list[str]:
    """Group multi-row FAQ entries into single text chunks."""
    chunks: list[str] = []
    current_q: str | None = None
    current_a: list[str] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        q = _clean(row[0])
        a = _clean(row[1])

        if q:
            if current_q and current_a:
                chunks.append(f"Вопрос: {current_q}\nОтвет: {' '.join(current_a)}")
            current_q = q
            current_a = [a] if a else []
        elif a and current_q:
            current_a.append(a)

    if current_q and current_a:
        chunks.append(f"Вопрос: {current_q}\nОтвет: {' '.join(current_a)}")

    return chunks


def parse_tariffs(ws) -> list[str]:
    """Create one text chunk per tariff column."""
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return []

    headers = [_clean(h) for h in rows[0]]
    chunks: list[str] = []

    for col_idx, name in enumerate(headers):
        if not name:
            continue
        lines = [name]
        for row in rows[1:]:
            val = _clean(row[col_idx]) if col_idx < len(row) else ""
            if val:
                lines.append(val)
        chunks.append("\n".join(lines))

    return chunks


def main() -> None:
    if not XLSX_PATH.exists():
        print(f"ERROR: file not found: {XLSX_PATH}", file=sys.stderr)
        sys.exit(1)

    print(f"Reading {XLSX_PATH.name} ...")
    wb = openpyxl.load_workbook(XLSX_PATH)

    chunks: list[str] = []

    if FAQ_SHEET in wb.sheetnames:
        faq = parse_faq(wb[FAQ_SHEET])
        print(f"  {FAQ_SHEET}: {len(faq)} entries")
        chunks.extend(faq)
    else:
        print(f"  WARNING: sheet '{FAQ_SHEET}' not found", file=sys.stderr)

    if TARIFF_SHEET in wb.sheetnames:
        tariffs = parse_tariffs(wb[TARIFF_SHEET])
        print(f"  {TARIFF_SHEET}: {len(tariffs)} tariff chunks")
        chunks.extend(tariffs)
    else:
        print(f"  WARNING: sheet '{TARIFF_SHEET}' not found", file=sys.stderr)

    if not chunks:
        print("ERROR: no chunks to ingest", file=sys.stderr)
        sys.exit(1)

    host = os.getenv("CHROMA_HTTP_HOST", "localhost")
    port = int(os.getenv("CHROMA_HTTP_PORT", "8000"))

    print(f"Connecting to ChromaDB at {host}:{port} ...")
    client = chromadb.HttpClient(host=host, port=port)

    try:
        client.delete_collection(COLLECTION_NAME)
        print(f"  Existing collection '{COLLECTION_NAME}' deleted.")
    except Exception:
        pass

    collection = client.create_collection(COLLECTION_NAME)
    collection.add(
        documents=chunks,
        ids=[f"chunk_{i}" for i in range(len(chunks))],
    )

    print(f"Done: {len(chunks)} chunks ingested into '{COLLECTION_NAME}'.")


if __name__ == "__main__":
    main()
